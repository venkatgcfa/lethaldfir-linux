"""
reports.xlsx_report
===================

LethalDFIR-branded XLSX workbook. Optional — requires ``openpyxl``.
Generated alongside the JSON / CSV / HTML reports.

Sheets
------
1. **Summary** — case header, host info, file inventory, severity counts.
2. **Findings** — every finding, severity-coloured, with evidence/metadata.
3. **Timeline** — full super-timeline, filterable, frozen header row.
4. **Login Records** — wtmp/btmp/utmp/lastlog events extracted from the
   case timeline (binary login analysis).
5. **Brute-Force Analysis** — top attacker IPs / targeted users derived
   from btmp + auth_log failed-login events; correlation against
   successful logins (compromise indicator).
6. **Tamper Detection** — file-size-vs-record-size integrity, EMPTY-record
   counts, and orphaned DEAD_PROCESS PIDs.

The colour palette matches the existing HTML report so deliverables look
consistent across formats.
"""

from __future__ import annotations

import os
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:                                                # pragma: no cover
    HAS_OPENPYXL = False

from .. import __brand__, __version__


# Regex matching XML-illegal control characters (everything except tab,
# newline, carriage-return in the C0 range, plus C1 surrogates).
# openpyxl raises IllegalCharacterError when these appear in cell values.
_ILLEGAL_XML_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
)


def _safe(value):
    """Sanitize a value for use in an XLSX cell.

    Strips control characters that are illegal in XML 1.0 (and therefore
    in XLSX) to prevent openpyxl ``IllegalCharacterError``.
    """
    if value is None:
        return ""
    if not isinstance(value, str):
        return value
    return _ILLEGAL_XML_RE.sub('', value)


# ---------------------------------------------------------------------------
# LethalDFIR purple palette (matches HTML report)
# ---------------------------------------------------------------------------
PURPLE_DARK   = "2D1B4E"
PURPLE_MID    = "5B2D8E"
PURPLE_LIGHT  = "F3EEFA"
PURPLE_HEADER = "3A1F5E"
WHITE         = "FFFFFF"

SEV_FILL = {
    "CRITICAL": "FF2E4E",
    "HIGH":     "FF7A39",
    "MEDIUM":   "F0B400",
    "LOW":      "4AD295",
    "INFO":     "5A9BF5",
}
RED_LIGHT    = "FFEBEE"
GREEN_LIGHT  = "E8F5E9"
YELLOW_LIGHT = "FFF8E1"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _set_col_width(ws, max_width: int = 60) -> None:
    """Auto-size columns to ~content width, capped at ``max_width``."""
    for col_cells in ws.columns:
        max_len = 10
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                v = "" if cell.value is None else str(cell.value)
            except Exception:
                v = ""
            if v:
                # only consider first line for width
                first = v.split("\n", 1)[0]
                max_len = max(max_len, min(len(first) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _ts(dt) -> str:
    if dt is None:
        return ""
    try:
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    except (AttributeError, ValueError):
        return str(dt)


def _styles():
    """Return the bundle of styles used across all sheets."""
    s = {
        "title":     Font(name="Arial", bold=True, color=PURPLE_DARK, size=14),
        "subtitle":  Font(name="Arial", bold=True, color=PURPLE_MID, size=11),
        "hdr_font":  Font(name="Arial", bold=True, color=WHITE, size=11),
        "hdr_fill":  PatternFill("solid", fgColor=PURPLE_HEADER),
        "hdr_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "data_font": Font(name="Arial", size=10),
        "data_align":Alignment(vertical="top", wrap_text=True),
        "alt_fill":  PatternFill("solid", fgColor=PURPLE_LIGHT),
        "warn_fill": PatternFill("solid", fgColor=RED_LIGHT),
        "ok_fill":   PatternFill("solid", fgColor=GREEN_LIGHT),
        "tip_fill":  PatternFill("solid", fgColor=YELLOW_LIGHT),
        "border":    Border(
            left=Side(style="thin", color="B39DDB"),
            right=Side(style="thin", color="B39DDB"),
            top=Side(style="thin", color="B39DDB"),
            bottom=Side(style="thin", color="B39DDB"),
        ),
    }
    return s


def _header_row(ws, row, ncols, styles):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = styles["hdr_font"]
        cell.fill = styles["hdr_fill"]
        cell.alignment = styles["hdr_align"]
        cell.border = styles["border"]


def _data_cell(ws, row, col, *, alt=False, warn=False, sev=None, styles=None):
    cell = ws.cell(row=row, column=col)
    cell.font = styles["data_font"]
    cell.alignment = styles["data_align"]
    cell.border = styles["border"]
    if sev and sev in SEV_FILL:
        cell.fill = PatternFill("solid", fgColor=SEV_FILL[sev])
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    elif warn:
        cell.fill = styles["warn_fill"]
    elif alt:
        cell.fill = styles["alt_fill"]


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def _build_summary_sheet(wb, case, styles) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = PURPLE_DARK

    ws.cell(row=1, column=1,
            value=f"{__brand__}  v{__version__}").font = styles["title"]
    ws.cell(row=2, column=1,
            value=f"Case: {case.case_name}").font = styles["subtitle"]
    ws.cell(row=3, column=1,
            value=f"Generated: {case.created_at.strftime('%Y-%m-%d %H:%M:%S UTC')}"
            ).font = Font(name="Arial", color=PURPLE_MID, size=10)
    ws.cell(row=4, column=1,
            value="https://LethalDFIR.com").font = Font(
                name="Arial", color=PURPLE_MID, size=10, italic=True)

    # ---- Host info ----
    row = 6
    ws.cell(row=row, column=1, value="Host Information").font = styles["subtitle"]
    hi = case.host_info or {}
    rows = [
        ("Hostname",   hi.get("hostname") or "(unknown)"),
        ("Distro",     hi.get("distro") or "(unknown)"),
        ("Kernel",     hi.get("kernel") or "(unknown)"),
        ("Evidence",   str(case.evidence_root)),
    ]
    for k, v in rows:
        row += 1
        ws.cell(row=row, column=1, value=k).font = Font(
            name="Arial", bold=True, size=10)
        ws.cell(row=row, column=2, value=_safe(str(v))).font = styles["data_font"]

    # ---- Host IP addresses ----
    ip_addrs = hi.get("ip_addresses") or []
    if ip_addrs:
        row += 1
        ws.cell(row=row, column=1, value="IP Addresses").font = Font(
            name="Arial", bold=True, size=10)
        ws.cell(row=row, column=2,
                value=", ".join(f"{e['ip']} ({e['iface']})" for e in ip_addrs)
                ).font = styles["data_font"]

        # Detail mini-table
        row += 2
        ws.cell(row=row, column=1, value="Network Interfaces").font = styles["subtitle"]
        row += 1
        ip_headers = ["Interface", "IP Address", "Source"]
        for i, h in enumerate(ip_headers, 1):
            ws.cell(row=row, column=i, value=h)
        _header_row(ws, row, len(ip_headers), styles)
        for i, e in enumerate(ip_addrs):
            row += 1
            ws.cell(row=row, column=1, value=e.get("iface", ""))
            ws.cell(row=row, column=2, value=e.get("ip", ""))
            ws.cell(row=row, column=3, value=e.get("source", ""))
            for c in range(1, len(ip_headers) + 1):
                _data_cell(ws, row, c, alt=(i % 2 == 0), styles=styles)

    # ---- Severity dashboard ----
    row += 2
    ws.cell(row=row, column=1, value="Findings by Severity").font = styles["subtitle"]
    row += 1
    counts = case.severity_counts()
    sev_headers = ["Severity", "Count"]
    for i, h in enumerate(sev_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _header_row(ws, row, len(sev_headers), styles)
    for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"):
        row += 1
        ws.cell(row=row, column=1, value=sev)
        ws.cell(row=row, column=2, value=counts.get(sev, 0))
        for c in range(1, 3):
            _data_cell(ws, row, c, styles=styles, sev=sev if c == 1 else None)

    # ---- Parser stats ----
    row += 2
    ws.cell(row=row, column=1, value="Parser Run Statistics").font = styles["subtitle"]
    row += 1
    headers = ["Parser", "Files", "Events", "Findings", "Errors"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _header_row(ws, row, len(headers), styles)

    for name in sorted(case.stats):
        s = case.stats[name]
        errs = s.get("errors") or []
        row += 1
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=s.get("files", 0))
        ws.cell(row=row, column=3, value=s.get("events", 0))
        ws.cell(row=row, column=4, value=s.get("findings", 0))
        ws.cell(row=row, column=5, value=len(errs))
        is_warn = bool(errs)
        for c in range(1, len(headers) + 1):
            _data_cell(ws, row, c, warn=is_warn, styles=styles)

    _set_col_width(ws)


def _build_findings_sheet(wb, case, styles) -> None:
    ws = wb.create_sheet("Findings")
    ws.sheet_properties.tabColor = "C62828"

    headers = ["Severity", "Category", "Title", "Description",
               "Artifact", "Time (UTC)", "Evidence", "Metadata"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _header_row(ws, 1, len(headers), styles)

    findings = case.findings_sorted()
    for idx, f in enumerate(findings):
        r = idx + 2
        ws.cell(row=r, column=1, value=_safe(f.severity))
        ws.cell(row=r, column=2, value=_safe(f.category))
        ws.cell(row=r, column=3, value=_safe(f.title))
        ws.cell(row=r, column=4, value=_safe(f.description))
        ws.cell(row=r, column=5, value=_safe(f.artifact))
        ws.cell(row=r, column=6, value=_ts(f.timestamp))
        ws.cell(row=r, column=7,
                value=_safe("\n".join(f.evidence) if f.evidence else ""))
        ws.cell(row=r, column=8,
                value=_safe("; ".join(f"{k}={v}" for k, v in (f.metadata or {}).items())))

        is_alt = idx % 2 == 0
        _data_cell(ws, r, 1, sev=f.severity, styles=styles)
        for c in range(2, len(headers) + 1):
            _data_cell(ws, r, c, alt=is_alt, styles=styles)

    if findings:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(findings)+1}"
    ws.freeze_panes = "A2"
    _set_col_width(ws)


def _build_timeline_sheet(wb, case, styles) -> None:
    ws = wb.create_sheet("Timeline")
    ws.sheet_properties.tabColor = PURPLE_MID

    headers = ["Time (UTC)", "Source", "Event Type", "User", "Host", "Description"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _header_row(ws, 1, len(headers), styles)

    events = case.sorted_events()
    for idx, e in enumerate(events):
        r = idx + 2
        ws.cell(row=r, column=1, value=_ts(e.timestamp))
        ws.cell(row=r, column=2, value=_safe(e.source))
        ws.cell(row=r, column=3, value=_safe(e.event_type))
        ws.cell(row=r, column=4, value=_safe(e.user or ""))
        ws.cell(row=r, column=5, value=_safe(e.host or ""))
        ws.cell(row=r, column=6, value=_safe((e.description or "")[:500]))
        is_alt = idx % 2 == 0
        for c in range(1, len(headers) + 1):
            _data_cell(ws, r, c, alt=is_alt, styles=styles)

    if events:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(events)+1}"
    ws.freeze_panes = "A2"
    _set_col_width(ws, max_width=80)


def _build_login_sheet(wb, case, styles) -> None:
    """Login-records sheet: wtmp/btmp/utmp/lastlog events."""
    ws = wb.create_sheet("Login Records")
    ws.sheet_properties.tabColor = PURPLE_MID

    sources = {"wtmp", "btmp", "utmp", "lastlog"}
    headers = ["Time (UTC)", "Source", "Event Type", "User",
               "Host / Remote", "Line / Terminal", "PID", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _header_row(ws, 1, len(headers), styles)

    rows_written = 0
    for idx, e in enumerate(ev for ev in case.sorted_events() if ev.source in sources):
        r = rows_written + 2
        rows_written += 1
        meta = e.metadata or {}
        ws.cell(row=r, column=1, value=_ts(e.timestamp))
        ws.cell(row=r, column=2, value=_safe(e.source))
        ws.cell(row=r, column=3, value=_safe(e.event_type))
        ws.cell(row=r, column=4, value=_safe(e.user or meta.get("user", "") or ""))
        ws.cell(row=r, column=5, value=_safe(e.host or meta.get("host", "") or ""))
        ws.cell(row=r, column=6, value=_safe(meta.get("line", "") or ""))
        ws.cell(row=r, column=7, value=_safe(str(meta.get("pid", "") or "")))
        notes = []
        if "type" in meta:
            notes.append(f"type={meta['type']}")
        if "uid" in meta:
            notes.append(f"uid={meta['uid']}")
        if "session" in meta:
            notes.append(f"session={meta['session']}")
        ws.cell(row=r, column=8, value=_safe(" ".join(notes)))

        is_alt = (rows_written - 1) % 2 == 0
        is_warn = e.source == "btmp" or e.event_type == "login_failed"
        for c in range(1, len(headers) + 1):
            _data_cell(ws, r, c, alt=is_alt, warn=is_warn, styles=styles)

    if rows_written:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{rows_written + 1}")
    ws.freeze_panes = "A2"
    _set_col_width(ws)


def _build_brute_force_sheet(wb, case, styles) -> None:
    """Top attacker IPs/users from btmp + auth_log failed-logins.

    Also flags compromise: IPs that appear in both failed and successful
    auth events.
    """
    ws = wb.create_sheet("Brute-Force Analysis")
    ws.sheet_properties.tabColor = "C62828"

    failed_ip_count: Counter[str] = Counter()
    failed_user_count: Counter[str] = Counter()
    failed_ip_user: dict[str, Counter] = defaultdict(Counter)
    success_ips: set[str] = set()
    failed_ips: set[str] = set()

    for e in case.events:
        meta = e.metadata or {}
        # btmp records (binary failed-logins) — line/host fields from utmp
        if e.source == "btmp":
            user = e.user or meta.get("user", "")
            ip   = e.host or meta.get("host", "")
            if ip:
                failed_ip_count[ip] += 1
                failed_ips.add(ip)
            if user:
                failed_user_count[user] += 1
            if ip and user:
                failed_ip_user[ip][user] += 1
        # auth.log style events — extract source IP from raw message
        elif e.source in ("auth.log", "secure"):
            etype = (e.event_type or "").lower()
            ip = meta.get("source_ip") or meta.get("addr") or ""
            user = e.user or meta.get("user", "")
            if "fail" in etype or "invalid" in etype:
                if ip:
                    failed_ip_count[ip] += 1
                    failed_ips.add(ip)
                if user:
                    failed_user_count[user] += 1
            elif "accept" in etype or "success" in etype:
                if ip:
                    success_ips.add(ip)

    total_failed = sum(failed_ip_count.values()) or sum(failed_user_count.values())
    compromised = failed_ips & success_ips

    # ---- Header ----
    ws.cell(row=1, column=1,
            value="Brute-Force Analysis").font = styles["title"]
    ws.cell(row=2, column=1,
            value=f"Total observed failed-login events: {total_failed}"
            ).font = styles["subtitle"]

    # ---- Compromise summary ----
    row = 4
    ws.cell(row=row, column=1, value="Compromise Indicator").font = styles["subtitle"]
    row += 1
    if compromised:
        ws.cell(row=row, column=1, value="WARNING")
        ws.cell(row=row, column=2,
                value=f"{len(compromised)} IP(s) appear in BOTH failed AND "
                      f"successful auth events: {', '.join(sorted(compromised))}")
        for c in range(1, 3):
            _data_cell(ws, row, c, warn=True, styles=styles)
    else:
        ws.cell(row=row, column=1, value="OK")
        ws.cell(row=row, column=2,
                value="No IPs observed in both failed and successful auth events.")

    # ---- Top IPs ----
    row += 2
    ws.cell(row=row, column=1, value="Top Attacker IPs (from btmp + auth_log)"
            ).font = styles["subtitle"]
    row += 1
    headers = ["Rank", "Source IP", "Failed Attempts", "% of Total", "Compromised?"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _header_row(ws, row, len(headers), styles)
    for i, (ip, n) in enumerate(failed_ip_count.most_common(25), 1):
        row += 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=ip)
        ws.cell(row=row, column=3, value=n)
        ws.cell(row=row, column=4,
                value=f"{n / total_failed * 100:.1f}%" if total_failed else "")
        is_comp = ip in compromised
        ws.cell(row=row, column=5, value="YES" if is_comp else "no")
        for c in range(1, len(headers) + 1):
            _data_cell(ws, row, c, alt=(i % 2 == 0), warn=is_comp, styles=styles)

    # ---- Top targeted users ----
    row += 2
    ws.cell(row=row, column=1,
            value="Top Targeted Usernames").font = styles["subtitle"]
    row += 1
    user_headers = ["Rank", "Username", "Failed Attempts", "% of Total"]
    for i, h in enumerate(user_headers, 1):
        ws.cell(row=row, column=i, value=h)
    _header_row(ws, row, len(user_headers), styles)
    user_total = sum(failed_user_count.values()) or 1
    for i, (user, n) in enumerate(failed_user_count.most_common(25), 1):
        row += 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=user)
        ws.cell(row=row, column=3, value=n)
        ws.cell(row=row, column=4, value=f"{n / user_total * 100:.1f}%")
        for c in range(1, len(user_headers) + 1):
            _data_cell(ws, row, c, alt=(i % 2 == 0), styles=styles)

    _set_col_width(ws)


def _build_user_accounts_sheet(wb, case, styles) -> None:
    """User accounts dump from /etc/passwd + /etc/shadow + /etc/group."""
    ws = wb.create_sheet("User Accounts")
    ws.sheet_properties.tabColor = PURPLE_MID

    headers = [
        "Username", "UID", "GID", "GECOS", "Home", "Shell",
        "Password Status", "Hash Algorithm",
        "Pwd Last Changed", "Account Expires",
        "Min Days", "Max Days", "Warn Days", "Inactive Days",
        "Never Expires", "Privileged?", "Service Acct?",
        "Interactive Shell?", "Supplementary Groups", "Anomalies",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _header_row(ws, 1, len(headers), styles)

    users = case.artifacts.get("local_users") or []
    users_sorted = sorted(
        users, key=lambda u: (u.get("uid") or 99999, u.get("name") or ""))

    for idx, u in enumerate(users_sorted):
        r = idx + 2
        grps = ",".join(u.get("groups") or [])
        anomalies = u.get("anomalies", "")
        is_alt = idx % 2 == 0
        is_warn = bool(anomalies)

        cells = [
            u.get("name", ""),
            u.get("uid", ""),
            u.get("gid", ""),
            u.get("gecos", ""),
            u.get("home", ""),
            u.get("shell", ""),
            u.get("password_status") or "",
            u.get("hash_algorithm") or "",
            u.get("password_last_changed") or "",
            u.get("account_expires") or "",
            u.get("min_days") or "",
            u.get("max_days") or "",
            u.get("warn_days") or "",
            u.get("inactive_days") or "",
            "yes" if u.get("never_expires") else "no",
            "yes" if u.get("is_privileged") else "no",
            "yes" if u.get("is_service_account") else "no",
            "yes" if u.get("has_interactive_shell") else "no",
            grps,
            anomalies,
        ]
        for c, v in enumerate(cells, 1):
            ws.cell(row=r, column=c, value=_safe(v))
            _data_cell(ws, r, c, alt=is_alt, warn=is_warn, styles=styles)

    if users_sorted:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{len(users_sorted) + 1}")
    ws.freeze_panes = "C2"
    _set_col_width(ws)



    """Surface findings tagged as tamper / anti-forensic indicators."""
    ws = wb.create_sheet("Tamper Detection")
    ws.sheet_properties.tabColor = "C62828"

    ws.cell(row=1, column=1,
            value="Tamper / Anti-Forensic Indicators").font = styles["title"]
    ws.cell(row=2, column=1,
            value="Findings that suggest log tampering, history wiping, or "
                  "active concealment.").font = Font(
                name="Arial", color=PURPLE_MID, italic=True, size=10)

    headers = ["Severity", "Category", "Title", "Artifact",
               "Description", "Evidence"]
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _header_row(ws, row, len(headers), styles)

    keywords = (
        "history -c", "unset histfile", "anti-forensic", "shred",
        "tamper", "rotate", "ld.so.preload", "rootkit", "auditd",
        "fail2ban", "selinux", "apparmor", "alignment",
    )
    relevant = []
    for f in case.findings_sorted():
        text = f"{f.title} {f.description} {f.category}".lower()
        if any(k in text for k in keywords):
            relevant.append(f)

    for idx, f in enumerate(relevant):
        r = row + 1 + idx
        ws.cell(row=r, column=1, value=_safe(f.severity))
        ws.cell(row=r, column=2, value=_safe(f.category))
        ws.cell(row=r, column=3, value=_safe(f.title))
        ws.cell(row=r, column=4, value=_safe(f.artifact))
        ws.cell(row=r, column=5, value=_safe(f.description))
        ws.cell(row=r, column=6,
                value=_safe("\n".join(f.evidence) if f.evidence else ""))
        _data_cell(ws, r, 1, sev=f.severity, styles=styles)
        for c in range(2, len(headers) + 1):
            _data_cell(ws, r, c, alt=(idx % 2 == 0), styles=styles)

    if not relevant:
        ws.cell(row=row + 1, column=1,
                value="No tamper / anti-forensic indicators detected.")

    _set_col_width(ws)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def write_xlsx(case, path) -> Path:
    """Build and save the LethalDFIR XLSX workbook for ``case``."""
    if not HAS_OPENPYXL:                                            # pragma: no cover
        raise RuntimeError(
            "openpyxl is not installed. Install it with: pip install openpyxl"
        )

    path = Path(path)
    wb = Workbook()
    styles = _styles()

    _build_summary_sheet(wb, case, styles)
    _build_findings_sheet(wb, case, styles)
    _build_user_accounts_sheet(wb, case, styles)
    _build_timeline_sheet(wb, case, styles)
    _build_login_sheet(wb, case, styles)
    _build_brute_force_sheet(wb, case, styles)
    _build_tamper_sheet(wb, case, styles)

    wb.save(path)
    return path
